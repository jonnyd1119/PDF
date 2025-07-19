import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import re
import os
import pdfplumber
from io import BytesIO
import tempfile
import shutil
import streamlit as st
import io

class CompletePlatform:
    def __init__(self):
        if 'configurations' not in st.session_state:
            st.session_state.configurations = {}
        if 'custom_upgrades' not in st.session_state:
            st.session_state.custom_upgrades = []
    
    def extract_text_from_pdf(self, pdf_file):
        try:
            with pdfplumber.open(pdf_file) as pdf:
                full_text = "\n".join(page.extract_text() or '' for page in pdf.pages)
            return full_text.lower()
        except Exception as e:
            st.error(f"Error reading PDF: {e}")
            return ""
    
    def analyze_excel_for_new_model(self, excel_file):
        try:
            wb = load_workbook(excel_file)
            first_sheet_name = wb.sheetnames[0]
            ws = wb[first_sheet_name]
            aircraft_model = first_sheet_name.replace(" FOR SALE", "").strip()
            
            field_analysis = {}
            potential_upgrades = {}
            avionics_section = {"start": None, "end": None, "rows": []}
            
            # First, find the AVIONICS UPGRADES section
            for row in range(1, min(101, ws.max_row + 1)):
                label_cell = ws.cell(row=row, column=12).value
                if label_cell and isinstance(label_cell, str):
                    label_upper = label_cell.upper().strip()
                    
                    # Check for AVIONICS UPGRADES header - be more flexible
                    if ("AVIONICS" in label_upper and "UPGRADE" in label_upper) or label_upper == "AVIONICS UPGRADES":
                        avionics_section["start"] = row + 1  # Start from next row
                        st.write(f"🔍 **Found AVIONICS UPGRADES section starting at row {row + 1}**")
                        
                        # Now find where it ends - look for the next section header
                        for end_row in range(row + 1, min(row + 50, ws.max_row + 1)):
                            end_cell = ws.cell(row=end_row, column=12).value
                            if end_cell and isinstance(end_cell, str):
                                end_upper = end_cell.upper().strip()
                                
                                # Common section headers that indicate end of avionics
                                if any(header in end_upper for header in ["INSPECTION", "INTERIOR", "EXTERIOR", "PAINT", "NOTES", "COMMENTS"]):
                                    avionics_section["end"] = end_row - 1
                                    st.write(f"🔍 **AVIONICS UPGRADES section ends at row {end_row - 1}**")
                                    break
                            elif not end_cell:
                                # Empty cell might indicate end of section
                                # Check if next few cells are also empty
                                all_empty = True
                                for check_row in range(end_row, min(end_row + 3, ws.max_row + 1)):
                                    if ws.cell(row=check_row, column=12).value:
                                        all_empty = False
                                        break
                                if all_empty:
                                    avionics_section["end"] = end_row - 1
                                    st.write(f"🔍 **AVIONICS UPGRADES section ends at row {end_row - 1} (empty rows)**")
                                    break
                        
                        # If we found start but no clear end, set a reasonable limit
                        if not avionics_section["end"]:
                            avionics_section["end"] = min(avionics_section["start"] + 25, ws.max_row)
                            st.write(f"🔍 **AVIONICS UPGRADES section ends at row {avionics_section['end']} (limit)**")
                        
                        break
            
            # Collect all avionics items
            if avionics_section["start"] and avionics_section["end"]:
                st.write(f"📋 **Collecting avionic items from rows {avionics_section['start']} to {avionics_section['end']}**")
                for row in range(avionics_section["start"], avionics_section["end"] + 1):
                    label_cell = ws.cell(row=row, column=12).value
                    if label_cell and isinstance(label_cell, str) and label_cell.strip():
                        item_name = label_cell.strip()
                        avionics_section["rows"].append({
                            "row": row,
                            "label": item_name,
                            "is_avionic": True
                        })
                        
                        # Also add to potential upgrades
                        upgrade_name = item_name.upper().replace(" ", "_").replace("-", "_")
                        keywords = [item_name.lower()]
                        if "-" in item_name:
                            keywords.append(item_name.lower().replace("-", " "))
                            keywords.append(item_name.lower().replace("-", ""))
                        
                        potential_upgrades[upgrade_name] = {
                            "row": row,
                            "label": item_name,
                            "keywords": keywords,
                            "is_yn": True,
                            "is_avionic": True
                        }
                        
                st.write(f"✅ **Found {len(avionics_section['rows'])} avionic items**")
            
            # First, scan the ITEM column (column 12) for items with default 'N' values
            item_column_upgrades = {}
            for row in range(1, min(101, ws.max_row + 1)):
                item_cell = ws.cell(row=row, column=12).value  # Column L (12)
                default_value_cell = ws.cell(row=row, column=13).value  # Column M (13)
                
                if item_cell and default_value_cell:
                    item_text = str(item_cell).strip()
                    default_value = str(default_value_cell).strip().upper()
                    
                    # If default value is 'N', this is likely an optional upgrade
                    if default_value == 'N':
                        # Skip generic terms
                        skip_terms = ['INSPECTION', 'INSPECTIONS', 'NOTES', 'COMMENTS', 'REMARKS']
                        if not any(skip == item_text.upper() for skip in skip_terms):
                            upgrade_name = item_text.upper().replace(" ", "_")
                            keywords = [item_text.lower()]
                            
                            # Add variations for common patterns
                            if "-" in item_text:
                                keywords.append(item_text.lower().replace("-", " "))
                                keywords.append(item_text.lower().replace("-", ""))
                            
                            item_column_upgrades[upgrade_name] = {
                                "row": row,
                                "label": item_text,
                                "keywords": keywords,
                                "is_yn": True,
                                "from_item_column": True
                            }
            
            # Then scan for Y/N fields by looking at existing broker columns
            yn_fields = {}
            for col in range(17, min(ws.max_column + 1, 50), 2):  # Check Y/N columns
                for row in range(1, min(101, ws.max_row + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip().upper() in ['Y', 'N']:
                        # Found a Y/N value, check what field it's for
                        label_cell = ws.cell(row=row, column=12).value
                        if label_cell and isinstance(label_cell, str):
                            label_text = label_cell.strip()
                            if label_text not in yn_fields:
                                yn_fields[label_text] = row
            
            # Now scan labels and incorporate findings
            for row in range(1, min(101, ws.max_row + 1)):
                label_cell = ws.cell(row=row, column=12).value
                
                if label_cell and isinstance(label_cell, str):
                    label_text = label_cell.upper().strip()
                    label_text_original = label_cell.strip()
                    
                    field_mappings = {
                        "year_model": ["YEAR MODEL", "YEAR & MODEL", "MODEL YEAR"],
                        "total_hours": ["TOTAL TIME SINCE NEW", "TOTAL HOURS", "HOURS SINCE NEW", "TTSN"],
                        "engine_overhaul": ["ENGINE TIME SINCE OVERHAUL", "ENGINE OVERHAUL", "TSOH"],
                        "engine_program": ["ENGINE PROGRAM", "ENGINE WARRANTY", "ENGINE PLAN"],
                        "apu_program": ["APU PROGRAM", "APU WARRANTY", "APU PLAN"],
                        "avionics_section": ["AVIONICS", "AVIONICS UPGRADES", "MISC AVIONICS"],
                        "number_of_seats": ["NUMBER OF SEATS", "SEATS", "SEATING"],
                        "seat_configuration": ["SEAT CONFIGURATION", "SEATING CONFIG", "INTERIOR CONFIG"],
                        "paint_exterior_year": ["PAINT EXTERIOR", "EXTERIOR PAINT", "PAINT YEAR", "PAINT COMPLETED"],
                        "interior_year": ["INTERIOR YEAR", "INTERIOR COMPLETED", "NEW INTERIOR", "INTERIOR REFURBISHMENT"]
                    }
                    
                    for field_name, patterns in field_mappings.items():
                        if any(pattern in label_text for pattern in patterns):
                            if field_name not in field_analysis:
                                field_analysis[field_name] = {
                                    "row": row,
                                    "label": label_cell,
                                    "confidence": "high"
                                }
                    
                    # Check if this is a Y/N field
                    if label_text_original in yn_fields:
                        # This is a Y/N upgrade field
                        upgrade_name = label_text.replace(" ", "_")
                        keywords = []
                        
                        # Skip generic INSPECTIONS row (but keep PREBUY INSPECTIONS)
                        if label_text == "INSPECTIONS":
                            continue
                        
                        # Generate keywords based on the label
                        if "BELTED LAV" in label_text:
                            keywords = ["belted lav", "belted lavatory", "bltd lav"]
                        elif "EXTERNAL LAV" in label_text:
                            keywords = ["external lav", "external lavatory", "ext lav"]
                        elif "GARMIN" in label_text:
                            keywords = [label_text_original.lower(), label_text_original.lower().replace("-", " ")]
                        elif "PREBUY" in label_text or "PRE-BUY" in label_text:
                            keywords = ["prebuy", "pre-buy", "pre buy", "prebuy inspection", "pre-buy inspection"]
                        else:
                            # Generic keywords from label
                            keywords = [label_text_original.lower()]
                        
                        potential_upgrades[upgrade_name] = {
                            "row": row,
                            "label": label_cell,
                            "keywords": keywords,
                            "is_yn": True
                        }
                    else:
                        # Check standard upgrade keywords
                        upgrade_keywords = [
                            "NXI", "G1000", "G3000", "G5000", "G-5000", "WIFI", "TCAS", "WAAS", "FANS", "DUAL FMS", "HF",
                            "AHRS", "FDR", "7.1 UPGRADE", "TCAS 7.1", "FLIGHT DATA RECORDER",
                            "SYNTHETIC VISION", "SVT", "GOGO", "IRIDIUM", "CPDLC", "MTOW", "MZFW",
                            "PREBUY INSPECTION", "PREBUY", "PRE-BUY", "DUAL UNS-1ESPW",
                            "APU", "AUXILIARY POWER UNIT", "BELTED LAV", "EXTERNAL LAV", "GARMIN"
                        ]
                        
                        for keyword in upgrade_keywords:
                            if keyword in label_text and label_text != "INSPECTIONS":  # Skip generic INSPECTIONS
                                upgrade_name = keyword.replace(" ", "_").upper()
                                if upgrade_name not in potential_upgrades:
                                    potential_upgrades[upgrade_name] = {
                                        "row": row,
                                        "label": label_cell,
                                        "keywords": [keyword.lower(), keyword.lower().replace("-", " ")]
                                    }
            
            # Merge item column upgrades with detected upgrades
            for upgrade_name, upgrade_info in item_column_upgrades.items():
                if upgrade_name not in potential_upgrades:
                    potential_upgrades[upgrade_name] = upgrade_info
            
            return {
                "aircraft_model": aircraft_model,
                "sheet_name": first_sheet_name,
                "fields": field_analysis,
                "upgrades": potential_upgrades,
                "yn_fields_detected": len(yn_fields),
                "item_upgrades_detected": len(item_column_upgrades),
                "avionics_section": avionics_section
            }
        
        except Exception as e:
            st.error(f"Error analyzing Excel: {e}")
            return None
    
    def create_configuration_interactive(self, analysis):
        st.subheader(f"🛩️ Configure: {analysis['aircraft_model']}")
        
        aircraft_model = st.text_input(
            "Aircraft Model Name:", 
            value=analysis['aircraft_model']
        )
        
        # Store avionics section info in session state
        if 'temp_avionics_section' not in st.session_state:
            st.session_state.temp_avionics_section = analysis.get('avionics_section', {})
        
        st.write("## 📋 Core Fields Mapping")
        field_mappings = {}
        
        core_fields = {
            "year_model": "Year Model",
            "total_hours": "Total Hours Since New", 
            "engine_overhaul": "Engine Time Since Overhaul",
            "engine_program": "Engine Program",
            "apu_program": "APU Program",
            "number_of_seats": "Number of Seats",
            "seat_configuration": "Seat Configuration",
            "paint_exterior_year": "Paint Exterior Year",
            "interior_year": "Interior Year"
        }
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("### Core Fields")
            for field_key, field_name in list(core_fields.items())[:5]:
                detected = analysis['fields'].get(field_key)
                if field_key == "apu_program":
                    default_row = detected['row'] if detected else 34
                else:
                    default_row = detected['row'] if detected else 20
                
                row_input = st.number_input(
                    f"{field_name}:",
                    min_value=1,
                    max_value=200,
                    value=default_row,
                    key=f"row_{field_key}"
                )
                field_mappings[field_key] = int(row_input)
        
        with col2:
            st.write("### More Fields")
            for field_key, field_name in list(core_fields.items())[5:]:
                detected = analysis['fields'].get(field_key)
                if field_key == "paint_exterior_year":
                    default_row = detected['row'] if detected else 60
                elif field_key == "interior_year":
                    default_row = detected['row'] if detected else 61
                else:
                    default_row = detected['row'] if detected else 30
                
                row_input = st.number_input(
                    f"{field_name}:",
                    min_value=1,
                    max_value=200,
                    value=default_row,
                    key=f"row_{field_key}"
                )
                field_mappings[field_key] = int(row_input)
        
        st.write("## 🔧 Upgrades Mapping")
        
        # Display avionics section info if found
        avionics_info = analysis.get('avionics_section', {})
        if avionics_info.get('rows'):
            st.success(f"✅ Detected AVIONICS UPGRADES section with {len(avionics_info['rows'])} items (rows {avionics_info['start']}-{avionics_info['end']})")
            st.info("These avionic items will be automatically copied with their formulas when adding new brokers")
        
        # Display detected Y/N fields
        if analysis.get('yn_fields_detected', 0) > 0:
            st.info(f"✅ Detected {analysis['yn_fields_detected']} Y/N upgrade fields from existing brokers!")
        
        if analysis.get('item_upgrades_detected', 0) > 0:
            st.info(f"✅ Detected {analysis['item_upgrades_detected']} optional upgrades from ITEM column with default 'N' values!")
        
        upgrade_mappings = {}
        
        # Initialize custom upgrades in session state
        if 'temp_custom_upgrades' not in st.session_state:
            st.session_state.temp_custom_upgrades = []
        
        # Add custom upgrade input section
        st.write("### Add Custom Upgrades")
        
        # Add multiple custom upgrades
        num_custom = st.number_input("Number of custom upgrades to add:", min_value=0, max_value=10, value=1)
        
        for i in range(num_custom):
            st.write(f"#### Custom Upgrade {i+1}")
            col1, col2, col3 = st.columns([2, 1, 2])
            with col1:
                custom_name = st.text_input(f"Upgrade Name {i+1}:", placeholder="e.g., GARMIN G-5000", key=f"custom_name_{i}")
            with col2:
                custom_row = st.number_input(f"Row {i+1}:", min_value=1, max_value=200, value=50, key=f"custom_row_{i}")
            with col3:
                custom_keywords = st.text_input(f"Keywords {i+1}:", placeholder="e.g., garmin g-5000, g5000", key=f"custom_keywords_{i}")
            
            if custom_name:
                upgrade_name = custom_name.upper().replace(" ", "_")
                if upgrade_name not in analysis['upgrades']:
                    analysis['upgrades'][upgrade_name] = {
                        "row": custom_row,
                        "label": custom_name,
                        "keywords": [kw.strip().lower() for kw in custom_keywords.split(",") if kw.strip()]
                    }
        
        st.write("### Detected Upgrades")
        if analysis['upgrades']:
            for upgrade_key, upgrade_info in analysis['upgrades'].items():
                col1, col2, col3 = st.columns([2, 1, 2])
                
                with col1:
                    include_upgrade = st.checkbox(
                        f"**{upgrade_key}**" + 
                        (" (Y/N field)" if upgrade_info.get('is_yn') else "") +
                        (" [from ITEM column]" if upgrade_info.get('from_item_column') else "") +
                        (" [AVIONIC]" if upgrade_info.get('is_avionic') else ""),
                        value=True,
                        key=f"upgrade_{upgrade_key}"
                    )
                
                with col2:
                    if include_upgrade:
                        row_input = st.number_input(
                            "Row:",
                            min_value=1,
                            max_value=200,
                            value=upgrade_info['row'],
                            key=f"upgrade_row_{upgrade_key}"
                        )
                
                with col3:
                    if include_upgrade:
                        keywords_input = st.text_input(
                            "Keywords:",
                            value=", ".join(upgrade_info['keywords']),
                            key=f"upgrade_keywords_{upgrade_key}"
                        )
                
                if include_upgrade:
                    upgrade_mappings[upgrade_key] = {
                        "keywords": [kw.strip().lower() for kw in keywords_input.split(",")],
                        "row": int(row_input)
                    }
        
        return aircraft_model, field_mappings, upgrade_mappings

    def generate_configuration(self, aircraft_model, field_mappings, upgrade_mappings):
        config = {
            aircraft_model: {
                "row_mappings": field_mappings,
                "upgrades": upgrade_mappings,
                "avionics_section": st.session_state.get('temp_avionics_section', {})
            }
        }
        return config
    
    def identify_aircraft_from_pdf(self, pdf_text):
        st.write(f"🔍 **Aircraft Model Debug**: Looking for matches in PDF")
        st.write(f"📋 **Configured models**: {list(st.session_state.configurations.keys())}")
        
        for model_name in st.session_state.configurations.keys():
            
            if model_name.lower() in pdf_text:
                st.write(f"✅ **Direct match found**: {model_name}")
                return model_name
            
            model_words = model_name.lower().split()
            
            if "excel" in model_name.lower():
                if "citation" in pdf_text and "excel" in pdf_text:
                    st.write(f"✅ **Citation Excel match found**: {model_name}")
                    return model_name
                elif "excel" in pdf_text:
                    st.write(f"✅ **Excel match found**: {model_name}")
                    return model_name
            
            if len(model_words) >= 2:
                matches = []
                for word in model_words:
                    if word not in ["-", "master", "for", "sale"] and len(word) > 2:
                        if word in pdf_text:
                            matches.append(word)
                
                important_words = [w for w in model_words if w not in ["-", "master", "for", "sale"] and len(w) > 2]
                if len(matches) >= len(important_words) * 0.6:
                    st.write(f"✅ **Partial match found**: {model_name} (matched {len(matches)}/{len(important_words)} words)")
                    return model_name
        
        st.write("❌ **No aircraft model matches found**")
        return None
    
    def extract_data_from_pdf(self, pdf_text, aircraft_model):
        config = st.session_state.configurations.get(aircraft_model)
        if not config or not isinstance(config, dict):
            st.error(f"❌ Invalid or missing configuration for {aircraft_model}")
            return {}
        
        extracted_data = {}
        row_mappings = config.get("row_mappings", {})
        
        st.write("🔍 **Starting PDF data extraction...**")
        st.write(f"📋 **Configured fields**: {list(row_mappings.keys())}")
        
        # Extract Serial Number
        serial_patterns = [
            r'serial\s+number[:\s]*([A-Z0-9\-]+)',
            r'sn[:\s]*([A-Z0-9\-]+)',
            r's/n[:\s]*([A-Z0-9\-]+)', 
            r'serial[:\s]*([A-Z0-9\-]+)',
            r'(\d{2,4}[A-Z]*\-?\d{3,4})',
            r'airframe[:\s]*([A-Z0-9\-]+)',
            r'aircraft[:\s]*([A-Z0-9\-]+)',
            r'msn[:\s]*([A-Z0-9\-]+)'
        ]
        
        for pattern in serial_patterns:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                serial_candidate = match.group(1).strip()
                if re.search(r'\d', serial_candidate) and len(serial_candidate) >= 3:
                    extracted_data["serial_number"] = serial_candidate
                    st.write(f"✅ **SERIAL NUMBER FOUND**: {serial_candidate}")
                    break
        
        # Extract other data
        year_patterns = [r'(19|20)\d{2}.*?(lear|citation|phenom|model)']
        for pattern in year_patterns:
            match = re.search(pattern, pdf_text)
            if match:
                year_match = re.search(r'(19|20)\d{2}', match.group(0))
                if year_match:
                    extracted_data["year_model"] = int(year_match.group(0))
                    break
        
        # Total Hours - Look specifically in ENGINE section
        total_hours_found = False
        
        # First try to find total hours in ENGINE section
        engines_section_match = re.search(r'engines?\s*[:\-\s]*([\s\S]{0,800}?)(?=\n\n|\navionics|\ninterior|\nexterior|$)', pdf_text, re.IGNORECASE)
        
        if engines_section_match:
            engines_text = engines_section_match.group(1)
            
            # Look for total time patterns within engine section
            engine_total_patterns = [
                r'engine\s*time\s*since\s*new[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'engine\s*ttsn[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'engine\s*total\s*time[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'total\s*time[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'(\d{1,2}[,\.]?\d{3})\s*hours?\s*total',
                r'(\d{1,2}[,\.]?\d{3})\s*total\s*hours?'
            ]
            
            for pattern in engine_total_patterns:
                match = re.search(pattern, engines_text, re.IGNORECASE)
                if match:
                    hours_str = match.group(1).replace(",", "").replace(".", "")
                    hours_value = int(hours_str)
                    extracted_data["total_hours"] = hours_value
                    st.write(f"✅ **TOTAL HOURS FOUND IN ENGINE SECTION**: {hours_value}")
                    total_hours_found = True
                    break
        
        # If not found in engine section, fall back to general patterns
        if not total_hours_found:
            total_hours_patterns = [
                r'(\d{1,2}[,\.]?\d{3})\s*airframe\s*hours\s*since\s*new',
                r'airframe\s*hours\s*since\s*new[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'total\s+time\s+since\s+new[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'(\d{1,2}[,\.]?\d{3})\s*hours\s*since\s*new',
                r'hours\s*since\s*new[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'hours\s*/?\s*new[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'ttsn[:\s]*(\d{1,2}[,\.]?\d{3})',
                r'total\s+hours[:\s]*(\d{1,2}[,\.]?\d{3})'
            ]
            
            for pattern in total_hours_patterns:
                match = re.search(pattern, pdf_text, re.IGNORECASE)
                if match:
                    hours_str = match.group(1).replace(",", "").replace(".", "")
                    hours_value = int(hours_str)
                    extracted_data["total_hours"] = hours_value
                    st.write(f"✅ **TOTAL HOURS FOUND**: {hours_value} (pattern: {pattern})")
                    break
        
        # Engine Overhaul - Look for various patterns
        engine_overhaul_patterns = [
            r'engine\s*time\s*since\s*overhaul[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'engine\s*tsoh[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'hours?\s*since\s*overhaul[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'time\s*since\s*overhaul[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'overhaul\s*hours?[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'tsoh[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'since\s*overhaul[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'(\d{1,2}[,\.]?\d{3})\s*hours?\s*since\s*overhaul',
            r'(\d{1,2}[,\.]?\d{3})\s*tsoh',
            r'(\d{1,2}[,\.]?\d{3})\s*since\s*overhaul',
            r'overhaul[:\s]*(\d{1,2}[,\.]?\d{3})',
            r'soh[:\s]*(\d{1,2}[,\.]?\d{3})'  # Sometimes abbreviated as SOH
        ]
        
        engine_overhaul_found = False
        
        # First try to find in ENGINE section if we have it
        if engines_section_match:
            engines_text = engines_section_match.group(1)
            
            for pattern in engine_overhaul_patterns:
                matches = re.findall(pattern, engines_text, re.IGNORECASE)
                if matches:
                    # Try each match to find a reasonable value
                    for match in matches:
                        hours_str = match.replace(",", "").replace(".", "")
                        hours_value = int(hours_str)
                        # Only accept reasonable overhaul hours (typically less than total hours)
                        if 100 <= hours_value <= 50000:  # Reasonable range
                            extracted_data["engine_overhaul"] = hours_value
                            st.write(f"✅ **ENGINE OVERHAUL FOUND IN ENGINE SECTION**: {hours_value}")
                            engine_overhaul_found = True
                            break
                if engine_overhaul_found:
                    break
        
        # If not found in engine section, try the whole document
        if not engine_overhaul_found:
            for pattern in engine_overhaul_patterns:
                matches = re.findall(pattern, pdf_text, re.IGNORECASE)
                if matches:
                    for match in matches:
                        hours_str = match.replace(",", "").replace(".", "")
                        hours_value = int(hours_str)
                        # Only accept reasonable overhaul hours
                        if 100 <= hours_value <= 50000:
                            extracted_data["engine_overhaul"] = hours_value
                            st.write(f"✅ **ENGINE OVERHAUL FOUND**: {hours_value}")
                            engine_overhaul_found = True
                            break
                if engine_overhaul_found:
                    break
        
        # If still not found, default to total hours
        if not engine_overhaul_found and "total_hours" in extracted_data:
            extracted_data["engine_overhaul"] = extracted_data["total_hours"]
            st.write(f"⚠️ **ENGINE OVERHAUL not found, defaulting to TOTAL HOURS**: {extracted_data['total_hours']}")
        
        # Engine Program - Look for specific phrases near "engine"
        engine_program_found = False
        
        # First, look specifically for engine program mentions
        engine_context_patterns = [
            r'engines?\s*(?:are\s*)?enrolled\s*in\s*([A-Za-z\s\-&]+?)(?:\s*full|\s*engine|\s*program|\.|\n|$)',
            r'both\s*engines?\s*(?:are\s*)?enrolled\s*in\s*([A-Za-z\s\-&]+?)(?:\s*full|\s*engine|\s*program|\.|\n|$)',
            r'engines?\s*(?:are\s*)?(?:on|under)\s*([A-Za-z\s\-&]+?)\s*(?:program|warranty|plan|coverage)',
            r'([A-Za-z\s\-&]+?)\s*full\s*engine\s*program',
            r'engines?\s*program[:\s]*([A-Za-z\s\-&]+?)(?:\n|$|\.)',
            r'engines?\s*warranty[:\s]*([A-Za-z\s\-&]+?)(?:\n|$|\.)',
            r'engines?\s*-\s*([A-Za-z\s\-&]+?)(?:\n|$|\.)',
            r'engine\s*maintenance[:\s]*([A-Za-z\s\-&]+?)(?:\n|$|\.)'
        ]
        
        # If we have an engine section, search there first
        if engines_section_match:
            engines_text = engines_section_match.group(1)
            
            for pattern in engine_context_patterns:
                match = re.search(pattern, engines_text, re.IGNORECASE)
                if match:
                    program = match.group(1).strip()
                    st.write(f"🔍 **Found potential engine program in engine section**: {program}")
                    
                    # Clean up the program name
                    program = program.strip().rstrip('.').rstrip(',')
                    
                    # Map to standard abbreviations
                    program_mapping = {
                        'jssi': 'JSSI',
                        'power advantage': 'PWR ADV',
                        'poweradvantage': 'PWR ADV',
                        'esp gold lite': 'ESP GOLD LITE',
                        'esp gold': 'ESP GOLD',
                        'esp': 'ESP',
                        'msp': 'MSP',
                        'csp': 'CSP',
                        'tap': 'TAP',
                        'tap blue': 'TAP BLUE',
                        'smart parts': 'SMART PARTS'
                    }
                    
                    program_lower = program.lower()
                    for key, value in program_mapping.items():
                        if key in program_lower:
                            extracted_data["engine_program"] = value
                            st.write(f"✅ **ENGINE PROGRAM FOUND**: {value}")
                            engine_program_found = True
                            break
                    
                    if engine_program_found:
                        break
        
        # If not found in engine section, search the whole document
        if not engine_program_found:
            for pattern in engine_context_patterns:
                match = re.search(pattern, pdf_text, re.IGNORECASE)
                if match:
                    program = match.group(1).strip()
                    
                    # Skip if it contains "avionics"
                    if "avionics" in program.lower():
                        continue
                    
                    st.write(f"🔍 **Found potential engine program**: {program}")
                    
                    # Clean up
                    program = program.strip().rstrip('.').rstrip(',')
                    
                    # Map to standard abbreviations
                    program_mapping = {
                        'jssi': 'JSSI',
                        'power advantage': 'PWR ADV',
                        'poweradvantage': 'PWR ADV',
                        'esp gold lite': 'ESP GOLD LITE',
                        'esp gold': 'ESP GOLD',
                        'esp': 'ESP',
                        'msp': 'MSP',
                        'csp': 'CSP',
                        'tap': 'TAP',
                        'tap blue': 'TAP BLUE',
                        'smart parts': 'SMART PARTS'
                    }
                    
                    program_lower = program.lower()
                    for key, value in program_mapping.items():
                        if key in program_lower:
                            extracted_data["engine_program"] = value
                            st.write(f"✅ **ENGINE PROGRAM FOUND**: {value}")
                            engine_program_found = True
                            break
                    
                    if engine_program_found:
                        break
                    
                    # If no mapping but seems valid, use it
                    if len(program) > 2 and len(program) < 20 and not any(skip in program.lower() for skip in ['avionics', 'triple', 'dual', 'honeywell']):
                        extracted_data["engine_program"] = program.upper()
                        engine_program_found = True
                        break
        
        # Number of Seats - Append " SEATS" to the number
        seats_patterns = [
            r'(\d+)\s*passenger\s*seat(?:ing|s)?',
            r'number\s*of\s*seats[:\s]*(\d+)',
            r'seats[:\s]*(\d+)',
            r'(\d+)\s*seat(?:s)?\s*(?:configuration|config)',
            r'seating\s*for\s*(\d+)',
            r'(\d+)\s*pax'
        ]
        
        for pattern in seats_patterns:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                seats = int(match.group(1))
                if 1 <= seats <= 20:
                    extracted_data["number_of_seats"] = f"{seats} SEATS"
                    st.write(f"✅ **NUMBER OF SEATS FOUND**: {seats} SEATS")
                    break
        
        # Seat Configuration
        config_patterns = [
            r'(?:seat(?:ing)?\s*)?config(?:uration)?[:\s]*([^\n]+)',
            r'cabin\s*config(?:uration)?[:\s]*([^\n]+)',
            r'interior\s*(?:features|has|with)[:\s]*([^\n]*(?:club|divan|forward facing|aft facing)[^\n]*)',
            r'(\d+\s*place\s*(?:club|divan)|forward\s*facing|aft\s*facing|center\s*club|double\s*club)[^\n]*'
        ]
        
        for pattern in config_patterns:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                seat_config = match.group(1).strip()
                if len(seat_config) > 5 and len(seat_config) < 200:
                    extracted_data["seat_configuration"] = seat_config
                    break
        
        # Paint Exterior Year
        paint_patterns = [
            r'paint(?:ed)?\s*(?:in\s*)?(\d{4})',
            r'exterior\s*paint(?:ed)?\s*(?:in\s*)?(\d{4})',
            r'(\d{4})\s*(?:exterior\s*)?paint',
            r'paint\s*completed[:\s]*(\d{4})',
            r'new\s*paint[:\s]*(\d{4})'
        ]
        
        for pattern in paint_patterns:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                year = int(match.group(1))
                if 1990 <= year <= 2030:
                    extracted_data["paint_exterior_year"] = year
                    break
        
        # Interior Year
        interior_patterns = [
            r'interior\s*(?:refurb(?:ished)?|completed|done|new)\s*(?:in\s*)?(\d{4})',
            r'(\d{4})\s*interior\s*(?:refurb|refresh|update)',
            r'new\s*interior[:\s]*(\d{4})',
            r'interior\s*year[:\s]*(\d{4})',
            r'refurb(?:ished)?\s*in\s*(\d{4})'
        ]
        
        for pattern in interior_patterns:
            match = re.search(pattern, pdf_text, re.IGNORECASE)
            if match:
                year = int(match.group(1))
                if 1990 <= year <= 2030:
                    extracted_data["interior_year"] = year
                    break
        
        # Extract Upgrades
        upgrades = config.get("upgrades", {})
        for upgrade_name, upgrade_config in upgrades.items():
            keywords = upgrade_config.get("keywords", [])
            found = False
            
            for keyword in keywords:
                if keyword and keyword in pdf_text:
                    extracted_data[f"upgrade_{upgrade_name}"] = "Y"
                    found = True
                    st.write(f"✅ **Upgrade found**: {upgrade_name} (keyword: {keyword})")
                    break
            
            if not found:
                extracted_data[f"upgrade_{upgrade_name}"] = "N"
        
        return extracted_data
    
    def find_broker_row(self, ws):
        """Find the row containing 'BROKER' label in column 12."""
        for row in range(1, min(50, ws.max_row + 1)):
            label_cell = ws.cell(row=row, column=12).value
            if label_cell and isinstance(label_cell, str):
                if "broker" in label_cell.lower():
                    return row
        return 5  # Default to row 5 if not found
    
    def shift_formulas_in_cell(self, cell, shift_amount):
        """
        Shift column references in Excel formulas while preserving absolute references.
        Absolute references (like $N45) are not shifted.
        """
        if not cell.value or not isinstance(cell.value, str):
            return
        
        if not cell.value.startswith('='):
            return
        
        formula = cell.value
        
        # Pattern to match column references, including absolute ones
        # This will match: A1, $A1, $A$1, A$1
        pattern = r'(\$?)([A-Z]+)(\$?)(\d+)'
        
        def replace_ref(match):
            dollar1 = match.group(1)  # $ before column (if any)
            col_letters = match.group(2)  # Column letters
            dollar2 = match.group(3)  # $ before row (if any)
            row_num = match.group(4)  # Row number
            
            # If column is absolute (has $ before it), do not shift
            if dollar1 == '$':
                return match.group(0)  # Return unchanged
            
            # Otherwise, shift the column
            col_num = 0
            for char in col_letters:
                col_num = col_num * 26 + (ord(char) - ord('A') + 1)
            
            new_col_num = col_num + shift_amount
            
            new_col_ref = ""
            while new_col_num > 0:
                new_col_num -= 1
                new_col_ref = chr(new_col_num % 26 + ord('A')) + new_col_ref
                new_col_num //= 26
            
            return f"{dollar1}{new_col_ref}{dollar2}{row_num}"
        
        new_formula = re.sub(pattern, replace_ref, formula)
        
        if new_formula != formula:
            cell.value = new_formula
    
    def find_insertion_point(self, excel_file, serial_number):
        """Find the correct column position to insert a new broker based on serial number order."""
        try:
            excel_content = excel_file.read()
            excel_file.seek(0)
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            wb = load_workbook(tmp_path)
            
            try:
                if '-' in serial_number:
                    new_serial_num = int(serial_number.split('-')[-1])
                    display_serial = serial_number.split('-')[-1]
                else:
                    new_serial_num = int(serial_number[-4:]) if len(serial_number) >= 4 else int(serial_number)
                    display_serial = serial_number[-4:] if len(serial_number) >= 4 else serial_number
            except:
                digits = ''.join(filter(str.isdigit, serial_number))
                new_serial_num = int(digits[-4:])
                display_serial = digits[-4:]
            
            st.write(f"🔍 **Looking for insertion point for serial**: {new_serial_num} (display: {display_serial})")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                serial_positions = []
                for col in range(16, ws.max_column + 1, 2):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and str(cell_value).strip():
                        try:
                            serial_str = str(cell_value).strip()
                            if serial_str.isdigit():
                                existing_serial = int(serial_str)
                            else:
                                if '-' in serial_str:
                                    existing_serial = int(serial_str.split('-')[-1])
                                else:
                                    digits = ''.join(filter(str.isdigit, serial_str))
                                    existing_serial = int(digits[-4:]) if len(digits) >= 4 else int(digits)
                            
                            serial_positions.append({
                                'column': col,
                                'serial': existing_serial,
                                'original': serial_str
                            })
                        except:
                            continue
                
                serial_positions.sort(key=lambda x: x['serial'])
                
                insert_col = None
                for pos in serial_positions:
                    if new_serial_num < pos['serial']:
                        insert_col = pos['column']
                        break
                
                if insert_col is None and serial_positions:
                    insert_col = serial_positions[-1]['column'] + 2
                elif insert_col is None:
                    insert_col = 16
                
                st.write(f"✅ **Insertion point found**: Column {insert_col}")
                st.write(f"📋 **Current serials**: {[p['original'] for p in serial_positions]}")
                
                os.unlink(tmp_path)
                return {
                    'column': insert_col,
                    'sheet': sheet_name,
                    'serial_positions': serial_positions,
                    'display_serial': display_serial
                }
            
            os.unlink(tmp_path)
            return None
            
        except Exception as e:
            st.error(f"Error finding insertion point: {e}")
            return None
    
    def find_broker_column(self, excel_file, serial_number, broker_name):
        try:
            excel_content = excel_file.read()
            excel_file.seek(0)
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            wb = load_workbook(tmp_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for col in range(16, ws.max_column + 1, 2):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and str(cell_value).strip() == serial_number:
                        st.write(f"✅ **Serial number {serial_number} found in existing column {col}**")
                        os.unlink(tmp_path)
                        return {
                            'column': col,
                            'sheet': sheet_name,
                            'broker': broker_name,
                            'matched_serial': serial_number,
                            'mode': 'update'
                        }
                
                insertion_info = self.find_insertion_point(excel_file, serial_number)
                if insertion_info:
                    insertion_info['broker'] = broker_name
                    insertion_info['matched_serial'] = serial_number
                    insertion_info['mode'] = 'insert'
                    os.unlink(tmp_path)
                    return insertion_info
            
            os.unlink(tmp_path)
            return None
            
        except Exception as e:
            st.error(f"Error finding broker column: {e}")
            return None
    
    def insert_new_row(self, excel_file, extracted_data, aircraft_model, insertion_info, serial_number):
        try:
            excel_file.seek(0)
            excel_content = excel_file.read()
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            backup_path = tmp_path.replace(".xlsx", "_BACKUP.xlsx")
            shutil.copy(tmp_path, backup_path)
            
            wb = load_workbook(tmp_path)
            ws = wb[insertion_info['sheet']]
            target_col = insertion_info['column']
            
            config = st.session_state.configurations[aircraft_model]
            row_mappings = config.get("row_mappings", {})
            updates = []
            
            st.write("🔍 **Row Insertion Debug**: Starting new row insertion")
            st.write(f"🔍 **Target column**: {target_col}")
            
            # Step 1: Shift existing columns with formula updating
            if insertion_info['serial_positions']:
                columns_to_shift = []
                for pos in insertion_info['serial_positions']:
                    if pos['column'] >= target_col:
                        columns_to_shift.append(pos['column'])
                
                columns_to_shift.sort(reverse=True)
                
                for source_col in columns_to_shift:
                    dest_col = source_col + 2
                    st.write(f"📋 **Shifting column {source_col} to {dest_col}**")
                    
                    for row in range(1, ws.max_row + 1):
                        source_cell = ws.cell(row=row, column=source_col)
                        dest_cell = ws.cell(row=row, column=dest_col)
                        
                        if source_cell.value is not None:
                            dest_cell.value = source_cell.value
                            
                            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                self.shift_formulas_in_cell(dest_cell, 2)
                            
                            try:
                                if source_cell.font:
                                    dest_cell.font = Font(
                                        name=source_cell.font.name,
                                        size=source_cell.font.size,
                                        bold=source_cell.font.bold,
                                        italic=source_cell.font.italic,
                                        color=source_cell.font.color
                                    )
                            except:
                                pass
                        
                        source_yn_cell = ws.cell(row=row, column=source_col + 1)
                        dest_yn_cell = ws.cell(row=row, column=dest_col + 1)
                        
                        if source_yn_cell.value is not None:
                            dest_yn_cell.value = source_yn_cell.value
                            
                            if isinstance(source_yn_cell.value, str) and source_yn_cell.value.startswith('='):
                                self.shift_formulas_in_cell(dest_yn_cell, 2)
                            
                            try:
                                if source_yn_cell.font:
                                    dest_yn_cell.font = Font(
                                        name=source_yn_cell.font.name,
                                        size=source_yn_cell.font.size,
                                        bold=source_yn_cell.font.bold,
                                        italic=source_yn_cell.font.italic,
                                        color=source_yn_cell.font.color
                                    )
                            except:
                                pass
                        
                        source_cell.value = None
                        source_yn_cell.value = None
            
            # Step 2: Find adjacent broker column to copy formulas from
            adjacent_col = None
            if target_col > 16:  # If not the first broker column
                # Use the column to the left as template
                adjacent_col = target_col - 2
            elif insertion_info['serial_positions']:  # If first position but others exist
                # Use the next column as template
                adjacent_col = target_col + 2
            
            # Step 3: Copy formulas from adjacent column (including avionics section)
            formulas_copied = 0
            avionics_copied = 0
            
            if adjacent_col and adjacent_col <= ws.max_column:
                st.write(f"📋 **Copying formulas from column {adjacent_col}**")
                
                # Get avionics section info if available
                avionics_section = config.get("avionics_section", {})
                avionics_start = avionics_section.get("start")
                avionics_end = avionics_section.get("end")
                
                for row in range(1, ws.max_row + 1):
                    # Skip certain rows that should have data, not formulas
                    if row == 1 or row == self.find_broker_row(ws):  # Serial number and broker rows
                        continue
                    
                    # Check if this row is in the avionics section
                    is_avionics_row = False
                    if avionics_start and avionics_end and avionics_start <= row <= avionics_end:
                        is_avionics_row = True
                    
                    # Check if this row is in our configured fields (these get data, not formulas)
                    is_data_row = any(row == row_num for row_num in row_mappings.values())
                    
                    # Also check if this is an upgrade row (but NOT in avionics section)
                    upgrades = config.get("upgrades", {})
                    is_upgrade_row = any(row == upgrade_config.get("row") for upgrade_config in upgrades.values()) and not is_avionics_row
                    
                    # Special handling for certain rows that should ALWAYS get formulas
                    # Get the actual row numbers from configuration
                    special_formula_rows = []
                    
                    # APU row (if configured)
                    apu_row = config.get("row_mappings", {}).get("apu_program")
                    if apu_row:
                        special_formula_rows.append(apu_row)
                    
                    # Check upgrades for rows that need formulas
                    for upgrade_name, upgrade_config in upgrades.items():
                        # Add any upgrade that typically has formulas
                        if any(keyword in upgrade_name for keyword in ["BELTED_LAV", "EXTERNAL_LAV", "DELIVERY_TO_THE_US", "DELIVERY", "APU"]):
                            upgrade_row = upgrade_config.get("row")
                            if upgrade_row:
                                special_formula_rows.append(upgrade_row)
                    
                    # Also add row 34 explicitly for APU if not already included
                    if 34 not in special_formula_rows:
                        special_formula_rows.append(34)
                    
                    # Add row 29 for DELIVERY TO THE US
                    if 29 not in special_formula_rows:
                        special_formula_rows.append(29)
                    
                    # Remove None values and ensure we have integers
                    special_formula_rows = [int(r) for r in special_formula_rows if r is not None]
                    
                    is_special_formula_row = row in special_formula_rows
                    
                    # For avionics rows, always copy formulas regardless of whether they're configured upgrades
                    if is_avionics_row:
                        # Check if this is just the header row (contains "AVIONICS" in the label)
                        label_cell = ws.cell(row=row, column=12).value
                        if label_cell and isinstance(label_cell, str) and "AVIONICS" in label_cell.upper() and "UPGRADE" in label_cell.upper():
                            # Skip the header row
                            continue
                        
                        # Copy main column
                        template_cell = ws.cell(row=row, column=adjacent_col)
                        new_cell = ws.cell(row=row, column=target_col)
                        
                        if template_cell.value:
                            new_cell.value = template_cell.value
                            
                            # If it's a formula, adjust references
                            if isinstance(template_cell.value, str) and template_cell.value.startswith('='):
                                col_shift = target_col - adjacent_col
                                self.shift_formulas_in_cell(new_cell, col_shift)
                            
                            # Copy formatting
                            try:
                                if template_cell.font:
                                    new_cell.font = Font(
                                        name=template_cell.font.name,
                                        size=template_cell.font.size,
                                        bold=template_cell.font.bold,
                                        italic=template_cell.font.italic,
                                        color=template_cell.font.color
                                    )
                            except:
                                pass
                            
                            avionics_copied += 1
                        
                        # Only handle Y/N column if there's actually an avionic item in this row
                        item_label = ws.cell(row=row, column=12).value
                        if item_label and isinstance(item_label, str) and item_label.strip():
                            # Copy Y/N column - set to N by default unless we found it in PDF
                            template_yn_cell = ws.cell(row=row, column=adjacent_col + 1)
                            new_yn_cell = ws.cell(row=row, column=target_col + 1)
                            
                            # Check if this avionic item was found in PDF
                            avionic_found = False
                            for upgrade_name, upgrade_config in upgrades.items():
                                if upgrade_config.get("row") == row:
                                    upgrade_key = f"upgrade_{upgrade_name}"
                                    if upgrade_key in extracted_data:
                                        new_yn_cell.value = extracted_data[upgrade_key]
                                        avionic_found = True
                                        break
                            
                            if not avionic_found:
                                # Default to N for avionics not found in PDF
                                new_yn_cell.value = "N"
                            
                            # Copy Y/N cell formatting
                            try:
                                if template_yn_cell.font:
                                    new_yn_cell.font = Font(
                                        name=template_yn_cell.font.name,
                                        size=template_yn_cell.font.size,
                                        bold=template_yn_cell.font.bold,
                                        italic=template_yn_cell.font.italic,
                                        color=template_yn_cell.font.color
                                    )
                            except:
                                pass
                    
                    elif not is_data_row and not is_upgrade_row or is_special_formula_row:
                        # Copy formula from adjacent column for non-data, non-upgrade rows OR special formula rows
                        template_cell = ws.cell(row=row, column=adjacent_col)
                        new_cell = ws.cell(row=row, column=target_col)
                        
                        if template_cell.value and isinstance(template_cell.value, str) and template_cell.value.startswith('='):
                            # Copy the formula
                            new_cell.value = template_cell.value
                            
                            # Adjust the formula references
                            col_shift = target_col - adjacent_col
                            self.shift_formulas_in_cell(new_cell, col_shift)
                            
                            # Copy formatting
                            try:
                                if template_cell.font:
                                    new_cell.font = Font(
                                        name=template_cell.font.name,
                                        size=template_cell.font.size,
                                        bold=template_cell.font.bold,
                                        italic=template_cell.font.italic,
                                        color=template_cell.font.color
                                    )
                            except:
                                pass
                            
                            formulas_copied += 1
                        
                        # Also check Y/N column for formulas
                        template_yn_cell = ws.cell(row=row, column=adjacent_col + 1)
                        new_yn_cell = ws.cell(row=row, column=target_col + 1)
                        
                        if template_yn_cell.value and isinstance(template_yn_cell.value, str) and template_yn_cell.value.startswith('='):
                            new_yn_cell.value = template_yn_cell.value
                            self.shift_formulas_in_cell(new_yn_cell, col_shift)
                            
                            try:
                                if template_yn_cell.font:
                                    new_yn_cell.font = Font(
                                        name=template_yn_cell.font.name,
                                        size=template_yn_cell.font.size,
                                        bold=template_yn_cell.font.bold,
                                        italic=template_yn_cell.font.italic,
                                        color=template_yn_cell.font.color
                                    )
                            except:
                                pass
                            
                            formulas_copied += 1
                
                if formulas_copied > 0:
                    st.write(f"✅ **Copied {formulas_copied} formulas from adjacent column**")
                
                if avionics_copied > 0:
                    st.write(f"✅ **Copied {avionics_copied} avionic items with formulas**")
            
            # Step 4: Add serial number (last 4 digits only) and broker name
            display_serial = insertion_info.get('display_serial', serial_number[-4:])
            
            ws.cell(row=1, column=target_col).value = display_serial
            updates.append(f"Serial Number - Row 1: {display_serial}")
            
            broker_row = self.find_broker_row(ws)
            if "broker_name" in extracted_data:
                broker_cell = ws.cell(row=broker_row, column=target_col)
                broker_cell.value = extracted_data["broker_name"].upper()  # Convert to uppercase
                updates.append(f"Broker Name - Row {broker_row}: {extracted_data['broker_name'].upper()}")
            else:
                broker_cell = ws.cell(row=broker_row, column=target_col)
                broker_cell.value = insertion_info.get('broker', 'Unknown Broker').upper()  # Convert to uppercase
                updates.append(f"Broker Name - Row {broker_row}: {insertion_info.get('broker', 'Unknown Broker').upper()}")
            
            # Step 5: Add extracted data
            for field, row_num in row_mappings.items():
                if field in extracted_data and row_num != 1 and row_num != broker_row:
                    new_value = extracted_data[field]
                    ws.cell(row=row_num, column=target_col).value = new_value
                    updates.append(f"{field} - Row {row_num}: {new_value}")
            
            # Step 6: Add upgrade data
            upgrades = config.get("upgrades", {})
            for upgrade_name, upgrade_config in upgrades.items():
                upgrade_key = f"upgrade_{upgrade_name}"
                if upgrade_key in extracted_data:
                    row_num = upgrade_config.get("row")
                    if row_num and row_num != 1:
                        yn_col = target_col + 1
                        cell = ws.cell(row=row_num, column=yn_col)
                        cell.value = extracted_data[upgrade_key]
                        try:
                            cell.font = Font(name="Calibri", size=11, color="FFFFFF")
                        except:
                            cell.font = Font(name="Calibri", size=11)
                        updates.append(f"{upgrade_name} - Row {row_num}: {extracted_data[upgrade_key]}")
            
            try:
                wb.calculation.calcMode = "automatic"
                st.write("✅ **Set calculation mode to automatic**")
            except:
                st.write("⚠️ **Could not set calculation mode**")
            
            wb.save(tmp_path)
            st.write("✅ **New row inserted successfully with formulas copied**")
            
            with open(tmp_path, "rb") as f:
                updated_excel_data = f.read()
            
            with open(backup_path, "rb") as f:
                backup_excel_data = f.read()
            
            os.unlink(tmp_path)
            os.unlink(backup_path)
            
            return updated_excel_data, backup_excel_data, updates
            
        except Exception as e:
            st.error(f"Error inserting new row: {e}")
            return None, None, []
    
    def update_excel(self, excel_file, extracted_data, aircraft_model, broker_info):
        try:
            # Add broker name to extracted data
            extracted_data["broker_name"] = broker_info['broker']
            
            if broker_info['mode'] == 'insert':
                return self.insert_new_row(excel_file, extracted_data, aircraft_model, broker_info, broker_info['matched_serial'])
            else:
                return self.update_existing_row(excel_file, extracted_data, aircraft_model, broker_info)
        except Exception as e:
            st.error(f"Error in update_excel: {e}")
            return None, None, []
    
    def update_existing_row(self, excel_file, extracted_data, aircraft_model, broker_info):
        try:
            excel_file.seek(0)
            excel_content = excel_file.read()
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            backup_path = tmp_path.replace(".xlsx", "_BACKUP.xlsx")
            shutil.copy(tmp_path, backup_path)
            
            wb = load_workbook(tmp_path)
            ws = wb[broker_info['sheet']]
            target_col = broker_info['column']
            
            config = st.session_state.configurations[aircraft_model]
            row_mappings = config.get("row_mappings", {})
            updates = []
            
            # Update broker name (convert to uppercase)
            broker_row = self.find_broker_row(ws)
            if "broker_name" in extracted_data:
                broker_cell = ws.cell(row=broker_row, column=target_col)
                broker_cell.value = extracted_data["broker_name"].upper()
                updates.append(f"Broker Name - Row {broker_row}: {extracted_data['broker_name'].upper()}")
            
            for field, row_num in row_mappings.items():
                if field in extracted_data:
                    if row_num == 1:
                        st.error(f"🚨 **PROTECTION**: Refusing to update Row 1 ({field}) - this is the serial number row!")
                        continue
                    
                    current_value = ws.cell(row=row_num, column=target_col).value
                    new_value = extracted_data[field]
                    
                    ws.cell(row=row_num, column=target_col).value = new_value
                    updates.append(f"{field} - Row {row_num}: {extracted_data[field]}")
            
            upgrades = config.get("upgrades", {})
            for upgrade_name, upgrade_config in upgrades.items():
                upgrade_key = f"upgrade_{upgrade_name}"
                if upgrade_key in extracted_data:
                    row_num = upgrade_config.get("row")
                    if row_num and row_num != 1:
                        yn_col = target_col + 1
                        cell = ws.cell(row=row_num, column=yn_col)
                        cell.value = extracted_data[upgrade_key]
                        try:
                            cell.font = Font(name="Calibri", size=11, color="FFFFFF")
                        except:
                            cell.font = Font(name="Calibri", size=11)
                        updates.append(f"{upgrade_name} - Row {row_num}: {extracted_data[upgrade_key]}")
            
            wb.save(tmp_path)
            
            with open(tmp_path, "rb") as f:
                updated_excel_data = f.read()
            
            with open(backup_path, "rb") as f:
                backup_excel_data = f.read()
            
            os.unlink(tmp_path)
            os.unlink(backup_path)
            
            return updated_excel_data, backup_excel_data, updates
            
        except Exception as e:
            st.error(f"Error updating existing row: {e}")
            return None, None, []

def main():
    st.set_page_config(page_title="Aircraft Data Platform", page_icon="✈️", layout="wide")
    
    platform = CompletePlatform()
    
    st.title("✈️ Complete Aircraft Data Management Platform")
    st.markdown("**Configure aircraft models and process broker data**")
    
    # Sidebar
    st.sidebar.title("📊 Status")
    
    if st.session_state.get('configurations'):
        st.sidebar.success(f"✅ {len(st.session_state.configurations)} Models Configured")
        for model in st.session_state.configurations.keys():
            st.sidebar.write(f"• {model}")
        
        # Debug: show configuration structure
        if st.sidebar.checkbox("Show Config Structure (Debug)"):
            for model, config in st.session_state.configurations.items():
                st.sidebar.write(f"Model: {model}")
                st.sidebar.write(f"Type: {type(config)}")
                if isinstance(config, dict):
                    st.sidebar.write(f"Keys: {list(config.keys())}")
        
        config_json = json.dumps(st.session_state.configurations, indent=2)
        st.sidebar.download_button(
            "📤 Export Configs",
            config_json,
            "aircraft_configurations.json",
            "application/json"
        )
        
        if st.sidebar.button("🗑️ Clear All"):
            st.session_state.configurations = {}
            st.session_state.custom_upgrades = []
            st.rerun()
    else:
        st.sidebar.warning("⚠️ No models configured")
    
    tab1, tab2 = st.tabs(["🚀 Quick Process", "🔧 Create New Model"])
    
    with tab1:
        st.header("🚀 Quick Aircraft Data Processing")
        
        # Load existing JSON
        st.subheader("📂 Load Existing Configuration")
        uploaded_json = st.file_uploader("Upload JSON Config", type="json")
        
        if uploaded_json:
            try:
                json_data = json.load(uploaded_json)
                st.write("**Found models:**")
                
                # Check if json_data is properly structured
                models_found = []
                for key, value in json_data.items():
                    if isinstance(value, dict) and "row_mappings" in value:
                        # This is a valid model configuration
                        models_found.append(key)
                        st.write(f"• {key}")
                    else:
                        st.warning(f"⚠️ Invalid configuration for key: {key}")
                
                if st.button("📥 Load Configuration"):
                    # Only load valid configurations
                    valid_configs = {k: v for k, v in json_data.items() 
                                   if isinstance(v, dict) and "row_mappings" in v}
                    
                    if valid_configs:
                        st.session_state.configurations = valid_configs
                        st.success(f"✅ Loaded {len(valid_configs)} configuration(s)!")
                        st.rerun()
                    else:
                        st.error("❌ No valid configurations found in the JSON file")
            except Exception as e:
                st.error(f"Error reading JSON: {e}")
        
        if st.session_state.get('configurations'):
            st.markdown("---")
            st.subheader("✈️ Process Aircraft Data")
            
            # Choose between single or multiple PDF mode
            process_mode = st.radio("Processing Mode:", ["Single PDF", "Multiple PDFs"], horizontal=True)
            
            if process_mode == "Single PDF":
                col1, col2 = st.columns(2)
                with col1:
                    serial_number = st.text_input("Serial Number:", placeholder="e.g., 0028", key="single_serial")
                with col2:
                    broker_name = st.text_input("Broker Name:", placeholder="e.g., FlyAlliance", key="single_broker")
                
                col1, col2 = st.columns(2)
                with col1:
                    pdf_file = st.file_uploader("Upload Broker PDF", type="pdf", key="single_pdf")
                with col2:
                    excel_file = st.file_uploader("Upload Excel Sheet", type="xlsx", key="single_excel")
                
                if pdf_file:
                    pdf_details = [{"pdf": pdf_file, "serial": serial_number, "broker": broker_name}]
                else:
                    pdf_details = []
            
            else:  # Multiple PDFs mode
                excel_file = st.file_uploader("Upload Excel Sheet", type="xlsx", key="multi_excel")
                
                st.write("### Add PDF Details")
                
                # Initialize session state for PDF details if not exists
                if 'pdf_entries' not in st.session_state:
                    st.session_state.pdf_entries = []
                
                # Add new PDF entry
                with st.expander("➕ Add New PDF Entry", expanded=True):
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        new_pdf = st.file_uploader("PDF File:", type="pdf", key="new_pdf")
                    with col2:
                        new_serial = st.text_input("Serial Number:", key="new_serial")
                    with col3:
                        new_broker = st.text_input("Broker Name:", key="new_broker")
                    
                    if st.button("➕ Add PDF", type="secondary"):
                        if new_pdf and new_serial and new_broker:
                            st.session_state.pdf_entries.append({
                                "pdf": new_pdf,
                                "serial": new_serial,
                                "broker": new_broker,
                                "name": new_pdf.name
                            })
                            st.success(f"Added {new_pdf.name}")
                            st.rerun()
                        else:
                            st.error("Please fill all fields before adding")
                
                # Display current PDF entries
                if st.session_state.pdf_entries:
                    st.write("### Current PDF Queue:")
                    for idx, entry in enumerate(st.session_state.pdf_entries):
                        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                        with col1:
                            st.write(f"📄 {entry['name']}")
                        with col2:
                            st.write(f"Serial: {entry['serial']}")
                        with col3:
                            st.write(f"Broker: {entry['broker']}")
                        with col4:
                            if st.button("❌", key=f"remove_{idx}"):
                                st.session_state.pdf_entries.pop(idx)
                                st.rerun()
                    
                    # Clear all button
                    if st.button("🗑️ Clear All PDFs"):
                        st.session_state.pdf_entries = []
                        st.rerun()
                
                pdf_details = st.session_state.pdf_entries
            
            if pdf_details and excel_file and all(d["serial"] and d["broker"] for d in pdf_details):
                if st.button("🚀 Process Aircraft Data", type="primary", key="process_btn"):
                    results = []
                    current_excel = excel_file
                    
                    for idx, detail in enumerate(pdf_details):
                        st.write(f"\n### Processing PDF {idx + 1} of {len(pdf_details)}: {detail.get('name', detail['pdf'].name)}")
                        
                        with st.spinner(f"Processing {detail.get('name', detail['pdf'].name)}..."):
                            pdf_text = platform.extract_text_from_pdf(detail['pdf'])
                            
                            if not pdf_text:
                                st.error(f"Could not extract text from {detail.get('name', detail['pdf'].name)}")
                                continue
                            
                            aircraft_model = platform.identify_aircraft_from_pdf(pdf_text)
                            
                            if not aircraft_model:
                                st.error(f"❌ Could not identify aircraft model in {detail.get('name', detail['pdf'].name)}")
                                st.info("Available: " + ", ".join(st.session_state.configurations.keys()))
                                continue
                            
                            st.success(f"✅ Identified: **{aircraft_model}**")
                            
                            extracted_data = platform.extract_data_from_pdf(pdf_text, aircraft_model)
                            
                            if extracted_data:
                                st.subheader("📊 Extracted Data")
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.write("**Core Fields:**")
                                    for key, value in extracted_data.items():
                                        if not key.startswith("upgrade_"):
                                            st.write(f"• **{key}**: {value}")
                                
                                with col2:
                                    st.write("**Upgrades:**")
                                    for key, value in extracted_data.items():
                                        if key.startswith("upgrade_"):
                                            upgrade_name = key.replace("upgrade_", "")
                                            icon = "✅" if value == "Y" else "❌"
                                            st.write(f"• {icon} **{upgrade_name}**: {value}")
                                
                                broker_info = platform.find_broker_column(current_excel, detail["serial"], detail["broker"])
                                
                                if not broker_info:
                                    st.error(f"❌ Could not find broker column or insertion point for {detail['serial']}")
                                    continue
                                
                                if broker_info['mode'] == 'update':
                                    st.success(f"✅ Found existing entry in Column {broker_info['column']} - will update")
                                else:
                                    st.success(f"✅ Will insert new row at Column {broker_info['column']}")
                                
                                updated_excel, backup_excel, updates = platform.update_excel(
                                    current_excel, extracted_data, aircraft_model, broker_info
                                )
                                
                                if updated_excel:
                                    mode_text = "updated" if broker_info['mode'] == 'update' else "inserted"
                                    st.success(f"✅ Excel {mode_text} successfully for {detail['serial']}!")
                                    
                                    # Save the result
                                    results.append({
                                        "serial": detail["serial"],
                                        "broker": detail["broker"],
                                        "pdf_name": detail.get('name', detail['pdf'].name),
                                        "updates": updates,
                                        "mode": broker_info['mode']
                                    })
                                    
                                    # Use the updated Excel for the next iteration
                                    if idx < len(pdf_details) - 1:
                                        # Create a file-like object for the next iteration
                                        current_excel = io.BytesIO(updated_excel)
                                        current_excel.name = "temp.xlsx"
                                else:
                                    st.error(f"❌ Failed to update Excel for {detail['serial']}")
                    
                    # Show summary and download
                    if results:
                        st.write("\n## 📊 Processing Summary")
                        for result in results:
                            st.write(f"• **{result['serial']}** ({result['broker']}): {result['mode']} - {len(result['updates'])} fields updated")
                        
                        st.write("\n## 📥 Download Results")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.download_button(
                                "📥 Download Updated Excel",
                                updated_excel,
                                f"UPDATED_MULTIPLE_{aircraft_model.replace(' ', '_')}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        with col2:
                            st.download_button(
                                "💾 Download Original Backup",
                                backup_excel,
                                f"ORIGINAL_BACKUP_{aircraft_model.replace(' ', '_')}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    # Clear PDF entries after successful processing
                    if process_mode == "Multiple PDFs" and results:
                        st.session_state.pdf_entries = []
            else:
                if not excel_file:
                    st.info("👆 Please upload an Excel file")
                elif not pdf_details:
                    st.info("👆 Please add at least one PDF with details")
                else:
                    st.info("👆 Please fill in all serial numbers and broker names")
        else:
            st.info("👆 Please load a configuration file first to enable data processing")
    
    with tab2:
        st.header("🔧 Create New Aircraft Model Configuration")
        
        st.subheader("🆕 Create New Configuration")
        excel_template = st.file_uploader("Upload Excel Template", type="xlsx")
        
        if excel_template:
            analysis = platform.analyze_excel_for_new_model(excel_template)
            
            if analysis:
                st.success(f"✅ Detected: **{analysis['aircraft_model']}**")
                
                aircraft_model, field_mappings, upgrade_mappings = platform.create_configuration_interactive(analysis)
                
                if st.button("Save Configuration", type="primary"):
                    config = platform.generate_configuration(aircraft_model, field_mappings, upgrade_mappings)
                    st.session_state.configurations.update(config)
                    st.success(f"✅ {aircraft_model} configured successfully!")
                    
                    # Clear temporary custom upgrades
                    if 'temp_custom_upgrades' in st.session_state:
                        del st.session_state.temp_custom_upgrades
                    
                    config_json = json.dumps(config, indent=2)
                    st.download_button(
                        "📥 Download Configuration File",
                        config_json,
                        f"{aircraft_model.replace(' ', '_')}_config.json",
                        "application/json",
                        key="download_new_config"
                    )
                    st.info("💡 **Important:** Download and save this configuration file! You'll need it for the Quick Process tab.")
                    if 'custom_upgrades' in st.session_state:
                        st.session_state.custom_upgrades = []

if __name__ == "__main__":
    main()
